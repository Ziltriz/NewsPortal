import openpyxl
import xlsxwriter
from io import BytesIO
from django.contrib.gis.geos import Point
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import RemarkablePlace, WeatherSummary
from .serializers import RemarkablePlaceSerializer, WeatherSummarySerializer
from rest_framework import status, permissions
from django.http import HttpResponse

class ImportPlacesFromXLSXView(APIView):
    permission_classes = [permissions.AllowAny]
    # permission_classes = [permissions.DjangoModelPermissionsOrAnonReadOnly]
    
    def get_queryset(self):
        return RemarkablePlace.objects.none()
    
    def post(self, request, *args, **kwargs):

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Не предоставлен файл'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
            
            places_to_create = []

            for row in worksheet.iter_rows(min_row=2, values_only=True):  
                name, lat, lon, rating = row

                if not all([name, lat, lon, rating]):
                    return Response({'error': f'Ошибка парсинга данных в стркое: {row}'}, status=status.HTTP_400_BAD_REQUEST)

                location = Point(float(lon), float(lat))

                places_to_create.append(
                    RemarkablePlace(name=name, location=location, rating=int(rating))
                )

            RemarkablePlace.objects.bulk_create(places_to_create)

            return Response({'message': f'{len(places_to_create)} мест импортировано успешно.'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ExportWeatherSummaryView(APIView):
    def get(self, request, *args, **kwargs):
        
        place_id = request.query_params.get('place')  
        date = request.query_params.get('date')       #

        # Фильтруем данные
        queryset = WeatherSummary.objects.all()

        if place_id:
            try:
                place = RemarkablePlace.objects.get(id=place_id)
                queryset = queryset.filter(place=place)
            except RemarkablePlace.DoesNotExist:
                return Response({'error': 'Place not found'}, status=status.HTTP_400_BAD_REQUEST)

        if date:
            queryset = queryset.filter(timestamp__date=date)

        # Создаем XLSX-файл
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Weather Summary')

        # Заголовки таблицы
        headers = ['Место', 'Дата', 'Температура (°C)', 'Влажность (%)', 'Давление (мм рт.ст.)', 'Направление ветра', 'Скорость ветра (м/с)']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)

        # Добавляем данные
        row_num = 1
        for summary in queryset:
            worksheet.write(row_num, 0, str(summary.place.name))
            worksheet.write(row_num, 1, summary.timestamp.strftime('%Y-%m-%d %H:%M'))
            worksheet.write(row_num, 2, summary.temperature)
            worksheet.write(row_num, 3, summary.humidity)
            worksheet.write(row_num, 4, summary.pressure)
            worksheet.write(row_num, 5, summary.wind_direction)
            worksheet.write(row_num, 6, summary.wind_speed)
            row_num += 1

        # Закрываем файл
        workbook.close()
        output.seek(0)

        # Создаем HTTP-ответ
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="weather_summary.xlsx"'
        return response